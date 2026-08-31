#!/usr/bin/env python
"""XLSX exporter for API cases (Roadmap 1.6) — renders api/cases.yaml into
``<Project>_v<N>_API_Cases.xlsx`` under the iteration's ``exports/`` dir.

Column contract (DATA_MODEL §7): ``api_case_id, module, operation_id, method,
endpoint, case_type, title, request.path_params, request.query,
request.headers, request.body, request.variables,
expected_response.status_code, expected_response.body_schema,
expected_response.body_includes, expected_response.body_assertions,
expected_response.derived_oracles``.

Byte-reproducibility contract (PRD §6): workbook document properties
(created/modified) are pinned and every ZIP entry of the resulting package is
rewritten with a fixed timestamp, so two runs over unchanged input produce
identical bytes. Versioning follows GLOSSARY: highest existing ``<N>`` in the
same exports/ dir plus one, starting at 1 — nothing is overwritten. Sources
are schema-gated through the shared registry; export refuses cases not in
``cases_valid``/``exported``.
"""

from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from _registry_lib import REPO_ROOT, RegistryError, _assert_safe_path, validate_path
from argus_core.parsing import load_yaml  # pyright: ignore[reportMissingImports]
from lint_test_design import lint_iteration
from openpyxl import Workbook, load_workbook  # pyright: ignore[reportMissingModuleSource]

PROJECT = REPO_ROOT.name
FIXED_DATE = (1980, 1, 1, 0, 0, 0)
FIXED_DATETIME = datetime(1980, 1, 1, 0, 0, 0)
FILENAME_PATTERN = re.compile(rf"^{PROJECT}_v(\d+)_API_Cases\.xlsx$")
CASES_READY = {"cases_valid", "exported"}

COLUMNS = [
    "api_case_id",
    "module",
    "operation_id",
    "method",
    "endpoint",
    "case_type",
    "title",
    "request.path_params",
    "request.query",
    "request.headers",
    "request.body",
    "request.variables",
    "expected_response.status_code",
    "expected_response.body_schema",
    "expected_response.body_includes",
    "expected_response.body_assertions",
    "expected_response.derived_oracles",
]


def cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def row_for(case: dict[str, Any]) -> list[str]:
    request = case.get("request") or {}
    expected = case.get("expected_response") or {}
    return [
        cell_value(case.get("api_case_id")),
        cell_value(case.get("module")),
        cell_value(case.get("operation_id")),
        cell_value(case.get("method")),
        cell_value(case.get("endpoint")),
        cell_value(case.get("case_type")),
        cell_value(case.get("title")),
        cell_value(request.get("path_params")),
        cell_value(request.get("query")),
        cell_value(request.get("headers")),
        cell_value(request.get("body")),
        cell_value(request.get("variables")),
        cell_value(expected.get("status_code")),
        cell_value(expected.get("body_schema")),
        cell_value(expected.get("body_includes")),
        cell_value(expected.get("body_assertions")),
        cell_value(expected.get("derived_oracles")),
    ]


def next_version(exports_dir: Path) -> int:
    highest = 0
    if exports_dir.is_dir():
        for existing in exports_dir.glob(f"{PROJECT}_v*_API_Cases.xlsx"):
            _assert_safe_path(existing, label="existing XLSX export")
            match = FILENAME_PATTERN.match(existing.name)
            if match:
                try:
                    version = int(match.group(1))
                except ValueError:
                    continue
                highest = max(highest, version)
    return highest + 1


def deterministic_xlsx_bytes(workbook: Workbook) -> bytes:
    workbook.properties.created = FIXED_DATETIME
    workbook.properties.modified = FIXED_DATETIME
    buffer = io.BytesIO()
    workbook.save(buffer)
    inner = zipfile.ZipFile(buffer)
    outer = io.BytesIO()
    with zipfile.ZipFile(outer, "w") as archive:
        for info in inner.infolist():
            pinned = zipfile.ZipInfo(info.filename, date_time=FIXED_DATE)
            pinned.compress_type = zipfile.ZIP_DEFLATED
            pinned.external_attr = info.external_attr
            payload = inner.read(info.filename)
            if info.filename == "docProps/core.xml":
                # openpyxl 在 save() 内部强制把 modified 改回当前时间，必须在
                # 最终 ZIP 层再次固定，否则跨秒两次导出的字节不同。
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>1980-01-01T00:00:00Z\g<2>",
                    payload,
                )
            archive.writestr(pinned, payload)
    return outer.getvalue()


def export(iteration_dir: Path) -> Path:
    _assert_safe_path(iteration_dir, label="iteration")
    if iteration_dir.is_symlink() or not iteration_dir.is_dir():
        raise RegistryError(f"iteration must be a safe directory: {iteration_dir}")
    cases_path = iteration_dir / "api" / "cases.yaml"
    if cases_path.is_symlink() or not cases_path.exists():
        raise RegistryError(f"missing source for export: {cases_path}")
    validate_path(cases_path)
    try:
        document = load_yaml(cases_path.read_bytes())
    except (OSError, UnicodeError, ValueError) as exc:
        raise RegistryError(f"source for export is not safely parseable: {cases_path}") from exc
    dependency_paths = (
        iteration_dir / "requirements.yaml",
        iteration_dir / "exemptions.yaml",
        iteration_dir / "api" / "spec.normalized.yaml",
    )
    if all(path.is_file() and not path.is_symlink() for path in dependency_paths):
        design_errors = [
            diagnostic
            for diagnostic in lint_iteration(iteration_dir, "api_cases")
            if diagnostic.severity == "error"
        ]
        if design_errors:
            detail = "; ".join(
                f"{diagnostic.rule_id} {diagnostic.location}: {diagnostic.message}"
                for diagnostic in design_errors[:5]
            )
            raise RegistryError(f"test-design lint failed before export: {detail}")
    status = document["status"]
    if status not in CASES_READY:
        raise RegistryError(
            f"api/cases.yaml status is {status!r}; export requires "
            f"{'/'.join(sorted(CASES_READY))} (PRD §4.4)"
        )

    workbook = Workbook()
    sheet = workbook.active
    if sheet is None:  # pragma: no cover - openpyxl always creates an active sheet
        raise RegistryError("openpyxl workbook has no active sheet")
    sheet.title = "API Cases"
    sheet.append(COLUMNS)
    for case in sorted(document["cases"], key=lambda c: c["api_case_id"]):
        sheet.append(row_for(case))

    exports_dir = iteration_dir / "exports"
    _assert_safe_path(exports_dir, label="exports directory")
    if exports_dir.is_symlink() or (exports_dir.exists() and not exports_dir.is_dir()):
        raise RegistryError(f"exports directory is not safe: {exports_dir}")
    exports_dir.mkdir(exist_ok=True)
    version = next_version(exports_dir)
    destination = exports_dir / f"{PROJECT}_v{version}_API_Cases.xlsx"
    if destination.is_symlink() or destination.exists():
        raise RegistryError(f"XLSX destination already exists: {destination}")
    # exports_dir is a checked directory and the destination is checked for
    # existing symlinks before this no-overwrite write.
    # pi-lens-ignore: python-path-traversal
    destination.write_bytes(deterministic_xlsx_bytes(workbook))
    return destination


def round_trip(path: Path) -> tuple[list[str], list[list[Any]]]:
    workbook = load_workbook(path)
    sheet = workbook.active
    if sheet is None:
        raise ValueError("workbook has no active sheet")
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(cell) for cell in rows[0]]
    return header, [list(row) for row in rows[1:]]


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=(__doc__ or "").splitlines()[0])
    parser.add_argument("iteration", type=Path, help="iterations/<id> directory")
    args = parser.parse_args(argv)

    iteration_dir = args.iteration if args.iteration.is_absolute() else REPO_ROOT / args.iteration
    if not iteration_dir.is_dir():
        print(f"error: iteration directory {iteration_dir} not found", file=sys.stderr)
        return 1
    try:
        destination = export(iteration_dir)
    except RegistryError as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    digest = hashlib.sha256(destination.read_bytes()).hexdigest()
    try:
        display_path = destination.relative_to(REPO_ROOT).as_posix()
    except ValueError:
        display_path = destination.as_posix()
    print(f"export_xlsx: wrote {display_path} (sha256 {digest})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
